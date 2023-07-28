from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox,Button
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd 
from openpyxl import Workbook
import pathlib
import tkinter as tk



# background="#06283D"
background="#020F29"
framebg="#EDEDED"
framefg="#06283D"



root=Tk()
root.title("Student Registration System")
root.geometry("1450x750+50+30")
# root.config(bg="black")
root.config(bg=background)



file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Date of Birth"
    sheet['D1']="Gender"
    sheet['E1']="EmailID"
    sheet['F1']="Contact NO"
    sheet['G1']="Country"
    sheet['H1']="State"
    sheet['I1']="City"
    sheet['J1']="Address"
    

    file.save('Student_data.xlsx')

######################ExitWindow######################

#Exit Window
def Exit():
    root.destroy()



#######################ShowImage#########################

def showimage():
    global filename
    global img
    # Assuming 'filename' contains the path to the image file
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetype=(("JPG File","*.jpg"),("PNG File","*.png"),("All files","*.txt")))

    img = Image.open(filename)
    resized_image = img.resize((135, 190))  # Provide width and height as a tuple
    tk_img = ImageTk.PhotoImage(resized_image)
    lbl.config(image=tk_img)
    lbl.img = tk_img





#####################Registration No.#######################
def registration_no():
    file=openpyxl.load_workbook('student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
    # print(max_row_value)

    try:
        Registration.set(max_row_value+1)

    except:    
        Registration.set("1")

####################Clear###########

def Clear():
    global img
    Name.set('')
    DOB.set('')
    radio.set(-1)
    EmailID.set('')
    ContactNo.set('')
    Country.set("Select Country")
    State.set("Select State")
    City.set("Select City")
    Address.set('')
    
    registration_no()

    saveButton.config(state='normal')

    img1=PhotoImage(file='images/uploadimage.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

##############Save#########
def save():
    R1=Registration.get()
    N1=Name.get()
    C1=Country.get()
    S1=State.get()
    C2=City.get()
    try:        
        G1=Gender
    except:
        messagebox.showerror("error","Select Gender!")  
        return 
    
    D2=DOB.get()
    D1=Date.get()
    E1=EmailID.get()
    C3=ContactNo.get()
    A1=Address.get()

    if R1=="" or N1=="" or C1=="" or S1=="" or C2=="" or D2=="" or D1=="" or E1=="" or C3=="" or A1=="":
        messagebox.showerror("error","Some Data is missing!")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')  
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1) 
        sheet.cell(column=2,row=sheet.max_row,value=N1) 
        sheet.cell(column=3,row=sheet.max_row,value=D2) 
        sheet.cell(column=4,row=sheet.max_row,value=G1) 
        sheet.cell(column=5,row=sheet.max_row,value=E1) 
        sheet.cell(column=6,row=sheet.max_row,value=C3) 
        sheet.cell(column=7,row=sheet.max_row,value=C1) 
        sheet.cell(column=8,row=sheet.max_row,value=S1) 
        sheet.cell(column=9,row=sheet.max_row,value=C2) 
        sheet.cell(column=10,row=sheet.max_row,value=A1) 
        
        file.save(r'Student_data.xlsx')


        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","  Picture is not available!!!")
        messagebox.showinfo("info","Successfully data entered!!" )

        Clear() #clear entry box and image section

        registration_no()  #recheck registration no. and takes new no.      


###########Search##############

def search():

    text = Search.get()
    Clear()
    saveButton.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            # print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            # print(reg_no_position)
            # print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Inavalid","Invalid registration number!!")      

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value  
    x3=sheet.cell(row=int(reg_number),column=3).value        
    x4=sheet.cell(row=int(reg_number),column=4).value        
    x5=sheet.cell(row=int(reg_number),column=5).value        
    x6=sheet.cell(row=int(reg_number),column=6).value        
    x7=sheet.cell(row=int(reg_number),column=7).value        
    x8=sheet.cell(row=int(reg_number),column=8).value        
    x9=sheet.cell(row=int(reg_number),column=9).value        
    x10=sheet.cell(row=int(reg_number),column=10).value   
        

    # print(x1)  
    # print(x2) 
    # print(x3) 
    # print(x4) 
    # print(x5)  
    # print(x6)  
    # print(x7) 
    # print(x8) 
    # print(x9) 
    # print(x10) 

    Registration.set(x1)
    Name.set(x2)
    DOB.set(x3)
    
    if x4=='Female':
       R2.select() 
    else:
        R1.select()

    EmailID.set(x5)
    ContactNo.set(x6)
    Country.set(x7)
    State.set(x8)
    City.set(x9)
    Address.set(x10)  

    img = (Image.open("Student Images/"+str(x1)+".jpg")) 
    resized_image=img.resize((135,190))    
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
       



########Update############
def Update():
    
    R1=Registration.get()
    N1=Name.get()
    C1=Country.get()
    S1=State.get()
    C2=City.get()

    selection()
    G1=Gender
    
    D2=DOB.get()
    D1=Date.get()
    E1=EmailID.get()
    C3=ContactNo.get()
    A1=Address.get()

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

    sheet.cell(column=2,row=int(reg_number),value=N1)   
    sheet.cell(column=2,row=int(reg_number),value=N1)   
    sheet.cell(column=3,row=int(reg_number),value=D2)            
    sheet.cell(column=4,row=int(reg_number),value=G1)            
    sheet.cell(column=5,row=int(reg_number),value=E1)            
    sheet.cell(column=6,row=int(reg_number),value=C3)            
    sheet.cell(column=7,row=int(reg_number),value=C1)            
    sheet.cell(column=8,row=int(reg_number),value=S1)            
    sheet.cell(column=9,row=int(reg_number),value=C2)            
    sheet.cell(column=10,row=int(reg_number),value=A1)            
   

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update","Update Succesfully!!")

    Clear()
        
    


#gender
def selection():
    global Gender
    value=radio.get()
    if value==1:
        Gender="Male"     
    else: 
        Gender="Female"
        

    
#top frames
Label(root,text="Email: amrutagali12@gmail.com ",width=10,height=1,bg="#96DED1").pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#01133E",fg="#fff",font='arial 20 bold').pack(side=TOP,fill=X)




#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=16,bd=2,font="arial 20").place(x=1020,y=40)
imageicon3=PhotoImage(file="images/search.png")

Srch=tk.Button(root,text="Search",compound=LEFT,image=imageicon3,width=100,bg='skyblue',font="arial 13 bold",command=search)
Srch.pack()
Srch.place(x=1270,y=40)

imageicon4=PhotoImage(file="Images/update.png")
Update_button=Button(root,image=imageicon4,bg="#01133E",command=Update)
Update_button.place(x=110,y=30)




#Registration and Date
Label(root,text="Registration No:",font="arial 15",fg=framebg,bg=background).place(x=50,y=150)
Label(root,text="Date:",font="arial 15",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date=StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 12")
reg_entry.place(x=200,y=155)

registration_no()





today=date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 12")
date_entry.place(x=555,y=155)

Date.set(d1)

#Students Details
obj=LabelFrame(root,text="Student's Details",font=30,bd=2,width=1100,bg=framebg,fg="#0E013E",height=510)
obj.place(x=30,y=200)

Label(obj,text="Full Name:", font="arial 14", bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:", font="arial 14", bg=framebg,fg=framefg).place(x=400,y=50)
Label(obj,text="gender:", font="arial 14", bg=framebg,fg=framefg).place(x=720,y=50)
Label(obj,text="Email ID:", font="arial 14", bg=framebg,fg=framefg).place(x=30,y=150)
Label(obj,text="Contact No:", font="arial 14", bg=framebg,fg=framefg).place(x=500,y=150)
Label(obj,text="Country:", font="arial 14", bg=framebg,fg=framefg).place(x=30,y=250)
Label(obj,text="State:", font="arial 14", bg=framebg,fg=framefg).place(x=400,y=250)
Label(obj,text="City:", font="arial 14", bg=framebg,fg=framefg).place(x=720,y=250)
Label(obj,text="Address:", font="arial 14", bg=framebg,fg=framefg).place(x=30,y=350)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 15")
name_entry.place(x=130,y=50)

DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=17,font="arial 15")
dob_entry.place(x=520,y=50)

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg="#020F29",command=selection)
R1.place(x=820,y=50)
R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg="#020F29",command=selection)
R2.place(x=900,y=50)

EmailID=StringVar()
email_entry=Entry(obj,textvariable=EmailID,width=20,font="arial 15")
email_entry.place(x=130,y=150)

ContactNo=StringVar()
contact_entry=Entry(obj,textvariable=ContactNo,width=20,font="arial 15")
contact_entry.place(x=610,y=150)

Country=Combobox(obj,values=['India'],font="Robot 14",width=20,state="r")
Country.place(x=130,y=255)
Country.set("Select Country")

State=Combobox(obj,values=['AndraPradesh','Assam','Bihar','Chattisgarh','Goa','Gujarat','Haryana','Jharkand','Karnataka','Kerala','Maharashtra','Meghalaya','Nagaland','Panjab','Rajastan','TamilNadu',],font="Robot 14",width=20,state="r")
State.place(x=460,y=255)
State.set("Select State")

City=Combobox(obj,values=['Bangalore','Bagalkote','Chitradurga','Davanagere','Gadag','Hubballi','Kolar','Mangaluru',],font="Robot 14",width=20,state="r")
City.place(x=770,y=255)
City.set("Select City")

Address=StringVar()
Address_entry=Entry(obj,textvariable=Address,width=30,font="arial 14")
Address_entry.place(x=130,y=350)


#image
f=Frame(root,bd=3,bg="black",width=170,height=220,relief=GROOVE)
f.place(x=1200,y=130)

img=PhotoImage(file="images/uploadimage.png")
lbl=Label(root,bg="black",image=img)
lbl.place(x=1217,y=147)

#button

UploadButton=Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="#7FB3D5",command=showimage)
UploadButton.place(x=1185,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="#3498DB",command=save)
saveButton.place(x=1185,y=470)

ResetButton=Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="#1F618D",command=Clear)
ResetButton.place(x=1185,y=570)

ExitButton=Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="#1B4F72",command=Exit)
ExitButton.place(x=1185,y=670)





root.mainloop()